import os
import openpyxl
import six

from zipfile import BadZipFile


class excel_reader:
    """
    Reads an Excel file (.xls or .xlsx) as a list of lists; a wrapper
    over the OpenPyXL interface.

    For files with multiple data sheets, a "sheet_name" parameter
    should be specified, otherwise the first sheet is chosen.
    """

    def __init__(self, target, sheet_name = None):
        self.workbook = openpyxl.load_workbook(target,
                                               read_only = True)
        if sheet_name is not None:
            self.sheet = self.workbook[sheet_name]
        else:
            sheets = self.workbook.get_sheet_names()
            self.sheet = self.workbook[sheets[0]]

        self.source = self.sheet.iter_rows()
        self.header = next(self.source)

    def __iter__(self):
        for row in self.source:
            values = [x.value for x in row]
            if not any(values):
                break
            yield values

    def close(self):
        # How to properly close an openpyxl sheet is still mostly
        # a matter of hearsay.
        self.wb._archive.close()
        del self.sheet
        del self.wb


class excel_writer:
    """
    Writes data to an Excel file (.xls or .xlsx.)  Opening an existing
    file while specifying a non-existent sheet will append that sheet
    to the file.

    If a sheet name is not specified the name defaults to "Sheet".
    """

    def __init__(self, filename, columns, sheet_name = None):
        self.filename = filename
        self.columns = columns

        if os.path.exists(filename):
            try:
                self.workbook = openpyxl.load_workbook(filename)
            except BadZipFile:
                raise IOError("Could not open Excel file to append sheet.")
            self.sheet = self.workbook.create_sheet(title = sheet_name)
        else:
            self.workbook = openpyxl.Workbook(write_only = True)
            if sheet_name and self.workbook.sheetnames:
                self.workbook.remove(self.workbook.sheetnames[0])
                self.sheet = self.workbook.create_sheet(sheet_name)

        self.row_index = 1

        self.write(self.columns)

    def write(self, row):
        if isinstance(row, dict):
            row = [row[x] for x in self.columns]

        for col_index, value in enumerate(row, start = 1):
            item = self.sheet.cell(row = self.row_index,
                                   column = col_index)
            item.value = value

        self.row_index += 1

    def close(self):
        self.wb.save(self.filename)


